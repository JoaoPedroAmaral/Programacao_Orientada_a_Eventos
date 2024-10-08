from customtkinter import *
from tkinter import *
from PIL import Image, ImageEnhance
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from collections import Counter
from tkinter import filedialog, messagebox
import os

largura_tela = 700
altura_tela = 640
imagem_icon_1 = Image.open(os.path.join(os.path.dirname(__file__), "icon_1.png"))


##--- Paleta de cores ---# 
cor_1 = '#566453'
cor_2 = '#212d1f'
cor_entry = '#080d06'
cor_texto = '#738270'
cor_frame = '#131d11'
cor_fundo = '#738270'
#________________________#


root = CTk()

class Funcoes():
    def abrir_imagem(self):
        try:
            self.panel = CTkLabel(self.frame_1, width=550, height=400, fg_color="transparent", text="")
            self.panel.place(relx=0.5, rely=0.48, anchor='center')

            self.open_imagem = filedialog.askopenfilename(title="Abrir Imagem")
            self.imagem = Image.open(self.open_imagem)
            self.imagem_origem = Image.open(self.open_imagem)

            w, h = self.imagem.size

            if w > h:
                self.img = CTkImage(light_image=self.imagem, dark_image=self.imagem, size=(490, 340))
            elif w == h:
                self.img = CTkImage(light_image=self.imagem, dark_image=self.imagem, size=(440, 340))
            else:
                self.img = CTkImage(light_image=self.imagem, dark_image=self.imagem, size=(300, 340))

            self.panel = CTkLabel(self.frame_1, image=self.img, text="")
            self.panel.place(relx=0.5, rely=0.48, anchor='center')

        except FileNotFoundError:
            print("Nenhum arquivo foi selecionado.")
        except IOError:
            print("Erro ao abrir a imagem.")
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")


    def mostrar_imagem(self):
        try:
            imagem_mostrar = self.new_imagem
            self.panel = CTkLabel(self.frame_1, width=550, height=400, fg_color="transparent", text="")
            self.panel.place(relx=0.5, rely=0.48, anchor='center')

            w, h = self.imagem.size

            if w > h:
                self.img = CTkImage(light_image= imagem_mostrar, dark_image= imagem_mostrar, size=(490, 340))
            elif w == h:
                self.img = CTkImage(light_image= imagem_mostrar, dark_image= imagem_mostrar, size=(440, 340))
            else:
                self.img = CTkImage(light_image= imagem_mostrar, dark_image= imagem_mostrar, size=(300, 340))
                
            self.panel = CTkLabel(self.frame_1, image=self.img, text="")
            self.panel.place(relx=0.5, rely=0.48, anchor='center')

        except AttributeError:
            print("Erro: A imagem não foi carregada corretamente.")
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")


    def trocar_imagem(self, tipo=0):
        try:
            if tipo == 1:
                self.imagem = self.imagem_CMYK
            else:
                self.imagem = self.new_imagem
            print("Imagem trocada")
        except AttributeError:
            print("Erro: A imagem especificada não foi carregada corretamente.")
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")


    def limpar_tudo(self):
        try:
            self.imagem = self.imagem_origem
            self.new_imagem = self.imagem_origem
            print("Alterações limpas")
            self.mostrar_imagem()
        except AttributeError as ae:
            print(f"Erro: {ae}. Certifique-se de que as imagens foram carregadas corretamente.")
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
        

    def baixar_imagem(self, ):
        try:
            self.new_imagem.save(f"Imagem_{self.indi}.jpeg")
            print(f"Imagem salva como: Imagem_{self.indi}.jpeg")
            self.indi += 1

        except AttributeError as ae:
            print(f"Erro de atributo: {ae}. Certifique-se de que as imagens foram carregadas corretamente.")
        except IOError as io:
            print(f"Erro ao salvar a imagem: {io}. Verifique se o caminho de destino é válido e você tem permissão para salvar arquivos.")
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")


    def mensagem_erro(self):
        self.sub_mensagem_de_erro()
        self.label = CTkLabel(self.sub_mensagem_erro, text="ERRO ao realizar esta opreação, tente novamente", font=('Arial', 12), text_color="Salmon")
        self.label.place(relx=0.5, rely=0.5, anchor='center')


#---------------------# Função Passar para excel #---------------------#
    def passar_para_excel(self):
        try:
            wb = Workbook()
            ws = wb.active

            for i in range(self.imagem.width):
                for j in range(self.imagem.height):
                    pixel = self.imagem.getpixel((i, j))
                    r, g, b = pixel
                    
                    fill_color = f'{r:02x}{g:02x}{b:02x}'
                    c = ws.cell(row=j+1, column=i+1)
                    c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

                    ws.column_dimensions[ws.cell(row=j+1, column=i+1).column_letter].width = 1.7
                    ws.row_dimensions[j+1].height = 9
            
            wb.save(f'test{self.indi}.xlsx')
            print(f"Arquivo test{self.indi}.xlsx salvo com sucesso")
            self.indi += 1

        except AttributeError as ae:
            print(f"Erro: {ae}. Certifique-se de que a imagem foi carregada corretamente.")
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")


#---------------------# Função Buscar Cor #---------------------#
    def buscar_RGB(self, cor_escolhida):
        try:
            for i in range(self.imagem.width):
                for j in range(self.imagem.height):
                    pixel = self.imagem.getpixel((i, j))
                    if isinstance(pixel, int):
                        r, g, b = pixel, pixel, pixel
                    elif len(pixel) == 3:
                        r, g, b = pixel
                    elif len(pixel) == 4:
                        r, g, b, a = pixel
                    else:
                        raise ValueError("Modo de cores não suportado")
                    
                    if (r, g, b) == cor_escolhida:
                        return True
            return False

        except AttributeError as ae:
            print(f"Erro: {ae}. Certifique-se de que a imagem foi carregada corretamente.")
            return False
        except ValueError as ve:
            print(f"Erro: {ve}")
            return False
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
            return False


    def buscar_cor_RGB(self):
        try:
            # Tenta obter e converter os valores de entrada
            self.R = int(self.entry_1.get())
            self.G = int(self.entry_2.get())
            self.B = int(self.entry_3.get())
            self.cor = (self.R, self.G, self.B)

            # Busca a cor na imagem
            if self.buscar_RGB(self.cor):
                print("Cor encontrada")
                self.existe = CTkLabel(self.sub_pagina, text="Cor encontrada", font=('Arial', 12), fg_color="green", bg_color="LimeGreen")
                self.existe.place(relx=0.5, rely=0.5, anchor='center')
            else:
                print("Cor não encontrada")
                self.nao_existe = CTkLabel(self.sub_pagina, text="Cor não encontrada", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
                self.nao_existe.place(relx=0.5, rely=0.5, anchor='center')

        except ValueError:
            print("Erro: Certifique-se de que todos os valores de entrada são números inteiros válidos.")
            self.erro_input = CTkLabel(self.sub_pagina, text="Erro: Entradas inválidas", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_input.place(relx=0.5, rely=0.5, anchor='center')
        except AttributeError as ae:
            print(f"Erro: {ae}. Certifique-se de que a imagem foi carregada corretamente.")
            self.erro_atributo = CTkLabel(self.sub_pagina, text="Erro: Imagem não carregada", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_atributo.place(relx=0.5, rely=0.5, anchor='center')
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
            self.erro_desconhecido = CTkLabel(self.sub_pagina, text="Erro inesperado", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_desconhecido.place(relx=0.5, rely=0.5, anchor='center')
        

    def buscar_cor_posicao(self):
        try:
            # Obtém as posições i e j a partir das entradas do usuário
            i = int(self.entry_1.get())
            j = int(self.entry_2.get())

            # Verifica se as coordenadas estão dentro dos limites da imagem
            if i < 0 or i >= self.imagem.width or j < 0 or j >= self.imagem.height:
                raise ValueError("Coordenadas fora dos limites da imagem")

            # Obtém a cor do pixel na posição especificada
            cor = self.imagem.getpixel((i, j))
            hex_color = '#{:02x}{:02x}{:02x}'.format(*cor[:3])

            # Determina a cor do texto para garantir visibilidade
            cor_texto = 'white' if sum(cor[:3]) / 3 < 128 else 'black'

            # Cria e posiciona a etiqueta com a cor encontrada
            self.existep = CTkLabel(self.sub_pagina, text=f'({cor[0]}, {cor[1]}, {cor[2]})', fg_color=hex_color, width=100, text_color=cor_texto)
            self.existep.place(relx=0.5, rely=0.5, anchor='center')

        except ValueError as ve:
            print(f"Erro: {ve}. Certifique-se de que as coordenadas são válidas e estão dentro dos limites da imagem.")
            self.erro_input = CTkLabel(self.sub_pagina, text="Erro: Coordenadas inválidas", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_input.place(relx=0.5, rely=0.5, anchor='center')
        except AttributeError as ae:
            print(f"Erro: {ae}. Certifique-se de que a imagem foi carregada corretamente.")
            self.erro_atributo = CTkLabel(self.sub_pagina, text="Erro: Imagem não carregada", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_atributo.place(relx=0.5, rely=0.5, anchor='center')
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
            self.erro_desconhecido = CTkLabel(self.sub_pagina, text="Erro inesperado", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_desconhecido.place(relx=0.5, rely=0.5, anchor='center')


#---------------------# Função Trocar Cor #---------------------#
    def trocar_cor(self):
        try:
            # Obtém e valida as cores de entrada
            r = int(self.R_entry_trocar.get())
            g = int(self.G_entry_trocar.get())
            b = int(self.B_entry_trocar.get())
            cor = (r, g, b)

            R_trocar = int(self.R_entry_trocar2.get())
            G_trocar = int(self.G_entry_trocar2.get())
            B_trocar = int(self.B_entry_trocar2.get())
            cor_trocar = (R_trocar, G_trocar, B_trocar)

            # Inicializa contagem de trocas e obtém dimensões da imagem
            trocas = 0
            w, h = self.imagem.size

            # Cria uma cópia da imagem para as modificações
            self.new_imagem = self.imagem.copy()

            # Percorre cada pixel da imagem
            for i in range(w):
                for j in range(h):
                    r, g, b = self.imagem.getpixel((i, j))
                    if (r, g, b) == cor:
                        self.new_imagem.putpixel((i, j), cor_trocar)
                        trocas += 1

            # Atualiza a interface
            self.sub_pagina.destroy()
            self.mostrar_imagem()
            print(f"Trocas realizadas: {trocas}")

        except ValueError:
            print("Erro: Certifique-se de que todos os valores de entrada são números inteiros válidos.")
            self.erro_input = CTkLabel(self.sub_pagina, text="Erro: Entradas inválidas", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_input.place(relx=0.5, rely=0.5, anchor='center')
        except AttributeError as ae:
            print(f"Erro: {ae}. Certifique-se de que a imagem foi carregada corretamente.")
            self.erro_atributo = CTkLabel(self.sub_pagina, text="Erro: Imagem não carregada", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_atributo.place(relx=0.5, rely=0.5, anchor='center')
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
            self.erro_desconhecido = CTkLabel(self.sub_pagina, text="Erro inesperado", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_desconhecido.place(relx=0.5, rely=0.5, anchor='center') 


    def listar_cores(self):
        try:
            # Conta as ocorrências de cada cor na imagem
            cor_da_imagem = Counter(self.imagem.getdata())
            # Obtém as 10 cores mais usadas
            cor_mais_usada = cor_da_imagem.most_common(10)
            
            print("10 cores mais usadas na imagem:")

            # Exibe as cores mais usadas
            for i, (cores, quantidade) in enumerate(cor_mais_usada):
                cor = cores
                hex_color = '#{:02x}{:02x}{:02x}'.format(*cor[:3])
                print(f"{cor}: {quantidade}")
                
                # Determina a cor do texto para garantir visibilidade
                cor_texto = 'white' if sum(cor[:3]) / 3 < 128 else 'black'
                
                k = (i + 2.5) / 13

                # Cria e posiciona a etiqueta com a cor encontrada
                self.label = CTkLabel(self.sub_pagina, text=f'({cor[0]}, {cor[1]}, {cor[2]})', fg_color=hex_color, width=100, text_color=cor_texto)
                self.label.place(relx=0.5, rely=k, anchor='center')

        except AttributeError as ae:
            print(f"Erro: {ae}. Certifique-se de que a imagem foi carregada corretamente.")
            self.erro_atributo = CTkLabel(self.sub_pagina, text="Erro: Imagem não carregada", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_atributo.place(relx=0.5, rely=0.5, anchor='center')
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
            self.erro_desconhecido = CTkLabel(self.sub_pagina, text="Erro inesperado", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_desconhecido.place(relx=0.5, rely=0.5, anchor='center')

        
#---------------------# Função Passar para CMYK #---------------------#
    def RGB_para_CMYK(self, rgb):
        # Função auxiliar para converter RGB para CMYK
        r, g, b = rgb
        if (r, g, b) == (0, 0, 0):
            # Caso especial de preto puro
            return 0, 0, 0, 1
        
        # Conversão para valores de 0 a 1
        r /= 255
        g /= 255
        b /= 255
        
        k = 1 - max(r, g, b)
        c = (1 - r - k) / (1 - k) if (1 - k) != 0 else 0
        m = (1 - g - k) / (1 - k) if (1 - k) != 0 else 0
        y = (1 - b - k) / (1 - k) if (1 - k) != 0 else 0
        
        return c, m, y, k

    def passar_CMYK(self):
        try:
            w, h = self.imagem.size
            self.imagem_CMYK = Image.new('CMYK', (w, h))
            
            for i in range(w):
                for j in range(h):
                    rgb = self.imagem.getpixel((i, j))
                    cor_cmyk = self.RGB_para_CMYK(rgb)
                    self.imagem_CMYK.putpixel((i, j), tuple(int(255 * value) for value in cor_cmyk))
            
            text1 = 1
            self.mostrar_imagem(text1)
            print("Transformado para CMYK")
        
        except AttributeError as ae:
            print(f"Erro: {ae}. Certifique-se de que a imagem foi carregada corretamente.")
            self.erro_atributo = CTkLabel(self.sub_pagina, text="Erro: Imagem não carregada", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_atributo.place(relx=0.5, rely=0.5, anchor='center')
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
            self.erro_desconhecido = CTkLabel(self.sub_pagina, text="Erro inesperado", font=('Arial', 12), fg_color="Red", bg_color="Salmon")
            self.erro_desconhecido.place(relx=0.5, rely=0.5, anchor='center')

    def RGB_para_CMYK(self, rgb):
        r = rgb[0] / 255.0
        g = rgb[1] / 255.0
        b = rgb[2] / 255.0
        c = 1.0 - r
        m = 1.0 - g
        y = 1.0 - b
        k = min(c, m, y)
        if k == 1.0:
            return (0, 0, 0, 1)
        c = (c - k) / (1.0 - k)
        m = (m - k) / (1.0 - k)
        y = (y - k) / (1.0 - k)
        return (c, m, y, k)


    def mostrar_imagem(self, tipo=0):
        self.panel = CTkLabel(self.frame_1, width=550, height=400, fg_color="transparent", text="")
        self.panel.place(relx=0.5, rely=0.48, anchor='center')

        w, h = self.imagem.size

        if tipo == 1:
            if w > h:
                self.img = CTkImage(light_image=self.imagem_CMYK, dark_image=self.imagem_CMYK, size=(490, 340))
            elif w == h:
                self.img = CTkImage(light_image=self.imagem_CMYK, dark_image=self.imagem_CMYK, size=(440, 340))
            else:
                self.img = CTkImage(light_image=self.imagem_CMYK, dark_image=self.imagem_CMYK, size=(300, 340))
        else:
            if w > h:
                self.img = CTkImage(light_image=self.new_imagem, dark_image=self.new_imagem, size=(490, 340))
            elif w == h:
                self.img = CTkImage(light_image=self.new_imagem, dark_image=self.new_imagem, size=(440, 340))
            else:
                self.img = CTkImage(light_image=self.new_imagem, dark_image=self.new_imagem, size=(300, 340))

        self.panel.configure(image=self.img)
        self.panel.place(relx=0.5, rely=0.48, anchor='center')


#---------------------# Função Passar para escala de cinza #---------------------#
    def passar_cinza_forma_1(self): # Media ponderada
        try:
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            
            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))
                    r, g, b = pixel
                    media = int((r + g + b) / 3)
                    self.new_imagem.putpixel((i, j), (media, media, media))
            
            self.mostrar_imagem()
            print("Transfornamado para Cinza - Media Ponderada")
        except Exception as e:
            self.mensagem_erro()


    def passar_cinza_forma_2(self): # Luminosidade com peso r=30 g= 20 b=15
        try:
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            
            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))
                    r, g, b = pixel
                    media = int((r * 0.30 + g * 0.20 + b * 0.15) / 0.65)
                    self.new_imagem.putpixel((i, j), (media, media, media))

            self.mostrar_imagem()
            print("Transfornamado para Cinza - Luminosidade")
        except Exception as e:
            self.mensagem_erro()
        

    def passar_cinza_forma_3(self): # Dessaturação
        try:
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            
            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))
                    r, g, b = pixel
                    max1 = max(pixel)
                    des = int((max1 - r + max1 - g + max1 - b) / 3)
                    self.new_imagem.putpixel((i, j), (des, des, des))

            
            self.mostrar_imagem()
            print("Transfornamado para Cinza - Dessaturação")
        except Exception as e:
            self.mensagem_erro()

    
    def passar_cinza_forma_4(self): # Decomposição de cores MÁXIMO
        try:
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            
            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))
                    max1 = max(pixel)
                    self.new_imagem.putpixel((i, j), (max1, max1, max1))
            self.mostrar_imagem()
            print("Transfornamado para Cinza - Decomposição de cores MÁXIMO")
        except Exception as e:
            self.mensagem_erro()
        

    def passar_cinza_forma_5(self): # Decomposição de cores MINIMO
        try:
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            
            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))
                    min1 = min(pixel)
                    self.new_imagem.putpixel((i, j), (min1, min1, min1))
            self.mostrar_imagem()
            print("Transfornamado para Cinza - Decomposição de cores MÍNIMO")
        except Exception as e:
            self.mensagem_erro()

    
#---------------------# Função Trocar Aumentar contraste cinza #---------------------#
    def aumentar_contraste_cinza_forma_1(self): # aumentar constraste usando normalização
        try:
        
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            
            cor_pixel = []
            nova_cor_pixel = []
            branco = 0

            for i in range(256):
                cor_pixel.append(0)
                nova_cor_pixel.append(0)

            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))
                    if pixel[1] != 255:
                        cor_pixel[pixel[1]] += 1
                    else:
                        branco += 1

            normalizacao = 0
            total = (w * h) - branco
            calculo_de_distribuicao_cumulativa = 0

            for i in range(256):
                if cor_pixel[i] != 0:
                    calculo_de_distribuicao_cumulativa = calculo_de_distribuicao_cumulativa + normalizacao
                    normalizacao = cor_pixel[i] / total

                    nova_cor_pixel[i] = int((normalizacao + calculo_de_distribuicao_cumulativa) * 255)

            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))

                    if pixel[1] == 255:
                        self.new_imagem.putpixel((i, j), (255, 255, 255))

                    else:
                        f_cor = nova_cor_pixel[pixel[1]]
                        self.new_imagem.putpixel((i, j), (f_cor, f_cor, f_cor))
                    
            self.mostrar_imagem()
            print("Aumentando Contraste")
        except Exception as e:
            self.mensagem_erro()


    def aumentar_contraste_cinza_forma_2(self):
        try:
        
            self.new_imagem = ImageEnhance.Contrast(self.imagem).enhance(1.5)

            self.mostrar_imagem()
            print("Aumentando Contraste")
        except Exception as e:
            self.mensagem_erro()


#---------------------# Função Aumentar contraste colorido #---------------------#
    def aumentar_contraste_colorido_forma_1(self):# aumentar o contraste de uma imagem colorida usando normalização e Distribuição de um Histograma para o contraste
        try:
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            
            cor_pixel_r = [0] * 256
            cor_pixel_g = [0] * 256
            cor_pixel_b = [0] * 256
            
            nova_cor_pixel_r = [0] * 256
            nova_cor_pixel_g = [0] * 256
            nova_cor_pixel_b = [0] * 256
            
            branco = 0

            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))
                    if pixel[0] != 255:
                        cor_pixel_r[pixel[0]] += 1
                    if pixel[1] != 255:
                        cor_pixel_g[pixel[1]] += 1
                    if pixel[2] != 255:
                        cor_pixel_b[pixel[2]] += 1
                    else:
                        branco += 1

            normalizacao_r = 0
            normalizacao_g = 0
            normalizacao_b = 0
            
            total = (w * h) - branco
            calculo_de_distribuicao_cumulativa_r = 0
            calculo_de_distribuicao_cumulativa_g = 0
            calculo_de_distribuicao_cumulativa_b = 0

            for i in range(256):
                if cor_pixel_r[i] != 0:
                    calculo_de_distribuicao_cumulativa_r = calculo_de_distribuicao_cumulativa_r + normalizacao_r
                    normalizacao_r = cor_pixel_r[i] / total
                    nova_cor_pixel_r[i] = int((normalizacao_r + calculo_de_distribuicao_cumulativa_r) * 255)

                if cor_pixel_g[i] != 0:
                    calculo_de_distribuicao_cumulativa_g = calculo_de_distribuicao_cumulativa_g + normalizacao_g
                    normalizacao_g = cor_pixel_g[i] / total
                    nova_cor_pixel_g[i] = int((normalizacao_g + calculo_de_distribuicao_cumulativa_g) * 255)
                    
                if cor_pixel_b[i] != 0:
                    calculo_de_distribuicao_cumulativa_b = calculo_de_distribuicao_cumulativa_b + normalizacao_b
                    normalizacao_b = cor_pixel_b[i] / total
                    nova_cor_pixel_b[i] = int((normalizacao_b + calculo_de_distribuicao_cumulativa_b) * 255)

            for i in range(w):
                for j in range(h):
                    pixel = self.imagem.getpixel((i, j))

                    if pixel[0] == 255 and pixel[1] == 255 and pixel[2] == 255:
                        self.new_imagem.putpixel((i, j), (255, 255, 255))

                    else:
                        f_cor_r = nova_cor_pixel_r[pixel[0]]
                        f_cor_g = nova_cor_pixel_g[pixel[1]]
                        f_cor_b = nova_cor_pixel_b[pixel[2]]
                        self.new_imagem.putpixel((i, j), (f_cor_r, f_cor_g, f_cor_b))
                        
            self.mostrar_imagem()
            print("Aumentando Contraste")
        except Exception as e:
            self.mensagem_erro()


    def aumentar_contraste_colorido_forma_2(self): # aumentar o contraste de uma imagem colorida
        try:
        
            self.new_imagem = ImageEnhance.Contrast(self.imagem).enhance(1.5)

            self.mostrar_imagem()
            print("Aumentando Contraste")
        except Exception as e:
            self.mensagem_erro()


#---------------------# Função Retirar um ou mais canais de cor de uma imagem #---------------------#
    def retirar_canal_de_cor(self, tipo, cor_retirada):
        try:
            w, h = self.imagem.size
            if tipo == 1:
                self.new_imagem = Image.new('RGB', (w, h))
                
                for i in range(w):
                    for j in range(h):
                        pixel = self.imagem.getpixel((i, j))
                        
                        if cor_retirada == 'R':
                            self.new_imagem.putpixel((i, j), (0, pixel[1], pixel[2]))
                        elif cor_retirada == 'G':
                            self.new_imagem.putpixel((i, j), (pixel[0], 0, pixel[2]))
                        elif cor_retirada == 'B':
                            self.new_imagem.putpixel((i, j), (pixel[0], pixel[1], 0))


                if cor_retirada == 'R':
                    print("Vermelho retirado")
                elif cor_retirada == 'G':
                    print("Verde retirado")
                elif cor_retirada == 'B':
                    print("Azul retirado")
                self.sub_pagina.destroy()
                self.mostrar_imagem()

            elif tipo == 2:
                self.new_imagem = self.imagem

                c, m, y, k = self.new_imagem.split()

                if cor_retirada == 'C':
                    c = c.point(lambda i: 0)            
                elif cor_retirada == 'M':
                    m = m.point(lambda i: 0)
                elif cor_retirada ==  'Y':
                    y = y.point(lambda i: 0)
                elif cor_retirada == 'K':
                    k = k.point(lambda i: 0)

                self.new_imagem = Image.merge('CMYK', (c, m, y, k))

                if cor_retirada == 'C':
                    print("Ciano retirado")
                
                elif cor_retirada == 'M':
                    print("Magenta retirado")

                elif cor_retirada ==  'Y':
                    print("Amarelo retirado")

                elif cor_retirada == 'K':
                    print("Preto retirado")

                self.sub_pagina.destroy()
                self.new_imagem = self.new_imagem.convert('RGB')
                self.mostrar_imagem()
                self.new_imagem = self.new_imagem.convert('CMYK')
                
        except Exception as e:
            self.mensagem_erro()


#---------------------# Função Implementação de filtros #---------------------#
    def filtro_1(self, fator):
        try:
            img_display =  self.imagem.copy()
            pixels = img_display.load()
            width, height = img_display.size
            
            def get_pixel(x, y):
                if 0 <= x < width and 0 <= y < height:
                    return pixels[x, y]
                return (0, 0, 0)

            kernel_size = fator
            kernel = [[1/(kernel_size**2)] * kernel_size for _ in range(kernel_size)]

            self.new_imagem = Image.new("RGB", img_display.size)
            
            new_pixels = self.new_imagem.load()

            for i in range(width):
                for j in range(height):
                    r_total, g_total, b_total = 0, 0, 0
                    for ki in range(-(kernel_size//2), kernel_size//2 + 1):
                        for kj in range(-(kernel_size//2), kernel_size//2 + 1):
                            r, g, b = get_pixel(i + ki, j + kj)
                            weight = kernel[ki + kernel_size//2][kj + kernel_size//2]
                            r_total += r * weight
                            g_total += g * weight
                            b_total += b * weight
                    new_pixels[i, j] = (int(r_total), int(g_total), int(b_total))
            
            self.sub_pagina.destroy()
            self.mostrar_imagem()
            print("Filtro 1")
        except Exception as e:
            self.mensagem_erro()


    def filtro_2(self):
        try:
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            

            img_display = self.imagem.copy()
            pixels = img_display.load()
            width, height = img_display.size
            
            def get_pixel(x, y):
                if 0 <= x < width and 0 <= y < height:
                    return pixels[x, y]
                return (0, 0, 0)

            kernel = [
                [-1, 0, 0],
                [0, 1, 0],
                [0, 0, 0]
            ]
            
            new_pixels = self.new_imagem.load()

            for i in range(width):
                for j in range(height):
                    r_edge, g_edge, b_edge = 0, 0, 0
                    for ki in range(-1, 2):
                        for kj in range(-1, 2):
                            r, g, b = get_pixel(i + ki, j + kj)
                            weight = kernel[ki + 1][kj + 1]
                            r_edge += r * weight
                            g_edge += g * weight
                            b_edge += b * weight
                    r_edge = min(max(0, r_edge), 255)
                    g_edge = min(max(0, g_edge), 255)
                    b_edge = min(max(0, b_edge), 255)
                    new_pixels[i, j] = (r_edge, g_edge, b_edge)
            
            self.mostrar_imagem()
            print("Filtro 2")
        except Exception as e:
            self.mensagem_erro()


    def filtro_3(self):
        try:
            w, h = self.imagem.size
            self.new_imagem = Image.new('RGB', (w, h))
            

            img_display = self.imagem.copy()
            pixels = img_display.load()
            width, height = img_display.size
            
            def get_pixel(x, y):
                if 0 <= x < width and 0 <= y < height:
                    return pixels[x, y]
                return (0, 0, 0)

            kernel = [
                [0, -1, 0],
                [-1, 5, -1],
                [0, -1, 0]
            ]
            
            new_pixels = self.new_imagem.load()

            for i in range(width):
                for j in range(height):
                    r_edge, g_edge, b_edge = 0, 0, 0
                    for ki in range(-1, 2):
                        for kj in range(-1, 2):
                            r, g, b = get_pixel(i + ki, j + kj)
                            weight = kernel[ki + 1][kj + 1]
                            r_edge += r * weight
                            g_edge += g * weight
                            b_edge += b * weight
                    r_edge = min(max(0, r_edge), 255)
                    g_edge = min(max(0, g_edge), 255)
                    b_edge = min(max(0, b_edge), 255)
                    new_pixels[i, j] = (r_edge, g_edge, b_edge)
        
            self.mostrar_imagem()
            print("Filtro 3")
        except Exception as e:
            self.mensagem_erro()


    def filtro_4(self):
        try:
            pixels = self.imagem.load()
            width, height = self.imagem.size
            
            def get_pixel(x, y):
                if 0 <= x < width and 0 <= y < height:
                    return pixels[x, y]
                return (0, 0, 0)

            sobel_x = [
                [-1, 0, 1],
                [-2, 0, 2],
                [-1, 0, 1]
            ]

            sobel_y = [
                [-1, -2, -1],
                [ 0,  0,  0],
                [ 1,  2,  1]
            ]

            self.new_imagem = Image.new("RGB",  self.imagem.size)
            
            new_pixels = self.new_imagem.load()

            for i in range(width):
                for j in range(height):
                    gx_r, gx_g, gx_b = 0, 0, 0
                    gy_r, gy_g, gy_b = 0, 0, 0
                    for ki in range(-1, 2):
                        for kj in range(-1, 2):
                            r, g, b = get_pixel(i + ki, j + kj)
                            gx_r += r * sobel_x[ki + 1][kj + 1]
                            gx_g += g * sobel_x[ki + 1][kj + 1]
                            gx_b += b * sobel_x[ki + 1][kj + 1]
                            gy_r += r * sobel_y[ki + 1][kj + 1]
                            gy_g += g * sobel_y[ki + 1][kj + 1]
                            gy_b += b * sobel_y[ki + 1][kj + 1]
                    edge_r = min(max(0, int((gx_r ** 2 + gy_r ** 2) ** 0.5)), 255)
                    edge_g = min(max(0, int((gx_g ** 2 + gy_g ** 2) ** 0.5)), 255)
                    edge_b = min(max(0, int((gx_b ** 2 + gy_b ** 2) ** 0.5)), 255)
                    new_pixels[i, j] = (edge_r, edge_g, edge_b)

            self.mostrar_imagem()
            print("Filtro 4")
        except Exception as e:
            self.mensagem_erro()



class Procetimentos():
    def Butoes(self):
        self.page_1()
        self.page_2()
        self.page_3()
        self.page_4()
        self.page_5()
        self.page_6()
        self.page_7()
        self.page_8()
        self.page_9()
        
    def page_1(self):
        label_1 = CTkLabel(master=self.tabview.tab(" 1 "), text="Transformar em excel", font=('Cooper Black', 15), text_color=cor_texto)
        label_1.place(relx=0.5, rely=0.1, anchor='center')


        self.btn_2_page_1 = CTkButton(master=self.tabview.tab(" 1 "), text="Transformar", command=self.passar_para_excel, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_1.place(relx=0.4, rely=0.5)


    def page_2(self):
        label_2 = CTkLabel(master=self.tabview.tab(" 2 "), text="Procurar Pixel", font=('Cooper Black', 15), text_color=cor_texto)
        label_2.place(relx=0.5, rely=0.1, anchor='center')

        self.btn_2_page_2 = CTkButton(master=self.tabview.tab(" 2 "), text="Por RGB", command=self.sub_page_1, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_2.place(relx=0.35, rely=0.6, anchor='center')

        self.btn_1_page_2 = CTkButton(master=self.tabview.tab(" 2 "), text="Por pixel", command=self.sub_page_2, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_1_page_2.place(relx=0.65, rely=0.6, anchor='center')


    def page_3(self):
        label_3 = CTkLabel(master=self.tabview.tab(" 3 "), text="Trocar cor", font=('Cooper Black', 15), text_color=cor_texto)
        label_3.place(relx=0.5, rely=0.1, anchor='center')

        self.btn_1_page_3 = CTkButton(master=self.tabview.tab(" 3 "), text="Trocar", command=self.sub_page_3, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_1_page_3.place(relx=0.35, rely=0.6, anchor='center')

        self.btn_2_page_3 = CTkButton(master=self.tabview.tab(" 3 "), text="Listar cores", command=self.sub_page_4, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_3.place(relx=0.65, rely=0.6, anchor='center')

        self.btn_3_page_3 = CTkButton(master=self.tabview.tab(" 3 "), text="", font=("arial", 12), width=50, height=25, command=self.baixar_imagem, 
                                      image=CTkImage(light_image=imagem_icon_1, dark_image=imagem_icon_1),corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_3_page_3.place(relx=0.02, rely=0.8)

        self.btn_4_page_3 = CTkButton(master=self.tabview.tab(" 3 "), text="Trocar", font=("arial", 12), width=50, height=25, command=self.trocar_imagem, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_4_page_3.place(relx=0.9, rely=0.8)

        self.btn_5_page_3 = CTkButton(master=self.tabview.tab(" 3 "), text="Limpar", font=("arial", 12), width=50, height=25, command=self.limpar_tudo, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_5_page_3.place(relx=0.9, rely=0.25)


    def page_4(self):
        label_4 = CTkLabel(master=self.tabview.tab(" 4 "), text="Transformar em CYMK", font=('Cooper Black', 15), text_color=cor_texto)
        label_4.place(relx=0.5, rely=0.1, anchor='center')


        self.btn_1_page_4 = CTkButton(master=self.tabview.tab(" 4 "), text="Transformar", command=self.passar_CMYK, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_1_page_4.place(relx=0.5, rely=0.6, anchor='center')


        self.btn_2_page_4 = CTkButton(master=self.tabview.tab(" 4 "), text="", font=("arial", 12), width=50, height=25, command= self.baixar_imagem, 
                                      image=CTkImage(light_image=imagem_icon_1, dark_image=imagem_icon_1),corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_4.place(relx=0.02, rely=0.8)

        self.btn_3_page_4 = CTkButton(master=self.tabview.tab(" 4 "), text="Trocar", font=("arial", 12), width=50, height=25, command=self.trocar_imagem, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_3_page_4.place(relx=0.9, rely=0.8)

        self.btn_3_page_4 = CTkButton(master=self.tabview.tab(" 4 "), text="Limpar", font=("arial", 12), width=50, height=25, command=self.limpar_tudo, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_3_page_4.place(relx=0.9, rely=0.25)


    def page_5(self):
        label_5 = CTkLabel(master=self.tabview.tab(" 5 "), text="Transformar em escala de Cinza", font=('Cooper Black', 15), text_color=cor_texto)
        label_5.place(relx=0.5, rely=0.1, anchor='center')

        self.btn_1_page_5 = CTkButton(master=self.tabview.tab(" 5 "), text="Forma 1", command=self.passar_cinza_forma_1, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_1_page_5.place(relx=0.25, rely=0.4, anchor='center')

        self.btn_2_page_5 = CTkButton(master=self.tabview.tab(" 5 "), text="Forma 2", command=self.passar_cinza_forma_2, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_5.place(relx=0.25, rely=0.7, anchor='center')

        self.btn_3_page_5 = CTkButton(master=self.tabview.tab(" 5 "), text="Forma 3", command=self.passar_cinza_forma_3, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_3_page_5.place(relx=0.5, rely=0.4, anchor='center')

        self.btn_4_page_5 = CTkButton(master=self.tabview.tab(" 5 "), text="Forma 4", command=self.passar_cinza_forma_4, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_4_page_5.place(relx=0.5, rely=0.7, anchor='center')

        self.btn_4_page_5 = CTkButton(master=self.tabview.tab(" 5 "), text="Forma 5", command=self.passar_cinza_forma_5, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_4_page_5.place(relx=0.75, rely=0.4, anchor='center')

        self.btn_5_page_5 = CTkButton(master=self.tabview.tab(" 5 "), text="", font=("arial", 12), width=50, height=25, command=self.baixar_imagem, 
                                      image=CTkImage(light_image=imagem_icon_1, dark_image=imagem_icon_1),corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_5_page_5.place(relx=0.02, rely=0.8)

        self.btn_6_page_5 = CTkButton(master=self.tabview.tab(" 5 "), text="Trocar", font=("arial", 12), width=50, height=25, command=self.trocar_imagem, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_6_page_5.place(relx=0.9, rely=0.8)

        self.btn_7_page_5 = CTkButton(master=self.tabview.tab(" 5 "), text="Limpar", font=("arial", 12), width=50, height=25, command=self.limpar_tudo, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_7_page_5.place(relx=0.9, rely=0.25)


    def page_6(self):
        label_6 = CTkLabel(master=self.tabview.tab(" 6 "), text="Aumentar contraste(Cinza)", font=('Cooper Black', 15), text_color=cor_texto)
        label_6.place(relx=0.5, rely=0.1, anchor='center')

        self.btn_1_page_6 = CTkButton(master=self.tabview.tab(" 6 "), text="Forma 1", command=self.aumentar_contraste_cinza_forma_1, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_1_page_6.place(relx=0.35, rely=0.6, anchor='center')

        self.btn_2_page_6 = CTkButton(master=self.tabview.tab(" 6 "), text="Forma 2", command=self.aumentar_contraste_cinza_forma_2, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_6.place(relx=0.65, rely=0.6, anchor='center')

        self.btn_3_page_6 = CTkButton(master=self.tabview.tab(" 6 "), text="", font=("arial", 12), width=50, height=25, command=self.baixar_imagem, 
                                      image=CTkImage(light_image=imagem_icon_1, dark_image=imagem_icon_1),corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_3_page_6.place(relx=0.02, rely=0.8)

        self.btn_4_page_6 = CTkButton(master=self.tabview.tab(" 6 "), text="Trocar", font=("arial", 12), width=50, height=25, command=self.trocar_imagem, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_4_page_6.place(relx=0.9, rely=0.8)

        self.btn_5_page_6 = CTkButton(master=self.tabview.tab(" 6 "), text="Limpar", font=("arial", 12), width=50, height=25, command=self.limpar_tudo, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_5_page_6.place(relx=0.9, rely=0.25)


    def page_7(self):
        label_8 = CTkLabel(master=self.tabview.tab(" 7 "), text="Aumentar contraste(Colorido)", font=('Cooper Black', 15), text_color=cor_texto)
        label_8.place(relx=0.5, rely=0.1, anchor='center')

        self.btn_1_page_7 = CTkButton(master=self.tabview.tab(" 7 "), text="Forma 1", command=self.aumentar_contraste_colorido_forma_1, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_1_page_7.place(relx=0.35, rely=0.6, anchor='center')

        self.btn_2_page_7 = CTkButton(master=self.tabview.tab(" 7 "), text="Forma 2", command=self.aumentar_contraste_colorido_forma_2, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_7.place(relx=0.65, rely=0.6, anchor='center')

        self.btn_3_page_7 = CTkButton(master=self.tabview.tab(" 7 "), text="", font=("arial", 12), width=50, height=25, command=self.baixar_imagem, 
                                      image=CTkImage(light_image=imagem_icon_1, dark_image=imagem_icon_1),corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_3_page_7.place(relx=0.02, rely=0.8)

        self.btn_4_page_7 = CTkButton(master=self.tabview.tab(" 7 "), text="Trocar", font=("arial", 12), width=50, height=25, command=self.trocar_imagem, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_4_page_7.place(relx=0.9, rely=0.8)

        self.btn_5_page_7 = CTkButton(master=self.tabview.tab(" 7 "), text="Limpar", font=("arial", 12), width=50, height=25, command=self.limpar_tudo, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_5_page_7.place(relx=0.9, rely=0.25)


    def page_8(self):
        label_8 = CTkLabel(master=self.tabview.tab(" 8 "), text="Remover canais de cor", font=('Cooper Black', 14), text_color=cor_texto)
        label_8.place(relx=0.5, rely=0.1, anchor='center')

        self.btn_1_page_8 = CTkButton(master=self.tabview.tab(" 8 "), text="Remover", command=self.sub_page_6, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_1_page_8.place(relx=0.5, rely=0.6, anchor='center')

        self.btn_2_page_8 = CTkButton(master=self.tabview.tab(" 8 "), text="", font=("arial", 12), width=50, height=25, command=self.baixar_imagem, 
                                      image=CTkImage(light_image=imagem_icon_1, dark_image=imagem_icon_1),corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_8.place(relx=0.02, rely=0.8)

        self.btn_3_page_8 = CTkButton(master=self.tabview.tab(" 8 "), text="Trocar", font=("arial", 12), width=50, height=25, command=self.trocar_imagem, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_3_page_8.place(relx=0.9, rely=0.8)

        self.btn_4_page_8 = CTkButton(master=self.tabview.tab(" 8 "), text="Limpar", font=("arial", 12), width=50, height=25, command=self.limpar_tudo, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_4_page_8.place(relx=0.9, rely=0.25)        


    def page_9(self):
        label_9 = CTkLabel(master=self.tabview.tab(" 9 "), text="Filtros Convolucionais", font=('Cooper Black', 15), text_color=cor_texto)
        label_9.place(relx=0.5, rely=0.1, anchor='center')

        self.btn_1_page_9 = CTkButton(master=self.tabview.tab(" 9 "), text="Filtro 1", command=self.sub_page_5, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_1_page_9.place(relx=0.35, rely=0.4, anchor='center')

        self.btn_2_page_9 = CTkButton(master=self.tabview.tab(" 9 "), text="Filtro 2", command=self.filtro_2, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_2_page_9.place(relx=0.35, rely=0.65, anchor='center')

        self.btn_3_page_9 = CTkButton(master=self.tabview.tab(" 9 "), text="Filtro 3", command= self.filtro_3, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_3_page_9.place(relx=0.65, rely=0.4, anchor='center')

        self.btn_4_page_9 = CTkButton(master=self.tabview.tab(" 9 "), text="Filtro 4", command= self.filtro_4, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_4_page_9.place(relx=0.65, rely=0.65, anchor='center')

        self.btn_5_page_9 = CTkButton(master=self.tabview.tab(" 9 "), text="", font=("arial", 12), width=50, height=25, command=self.baixar_imagem, 
                                      image=CTkImage(light_image=imagem_icon_1, dark_image=imagem_icon_1),corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_5_page_9.place(relx=0.02, rely=0.8)

        self.btn_6_page_9 = CTkButton(master=self.tabview.tab(" 9 "), text="Trocar", font=("arial", 12), width=50, height=25, command=self.trocar_imagem, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_6_page_9.place(relx=0.9, rely=0.8)

        self.btn_7_page_9 = CTkButton(master=self.tabview.tab(" 9 "), text="Limpar", font=("arial", 12), width=50, height=25, command=self.limpar_tudo, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_7_page_9.place(relx=0.9, rely=0.25)


    def sub_page_1(self):
        self.sub_Pagina(200, 290)

        self.entry_1 = CTkEntry(master=self.sub_pagina, width=50, font= ('consolas', 12), fg_color='transparent', text_color=cor_texto)
        self.entry_2 = CTkEntry(master=self.sub_pagina, width=50, font= ('consolas', 12), fg_color='transparent', text_color=cor_texto)
        self.entry_3 = CTkEntry(master=self.sub_pagina, width=50, font= ('consolas', 12), fg_color='transparent', text_color=cor_texto)

        self.entry_1.place(relx=0.5, rely=0.1, anchor='center')
        self.entry_2.place(relx=0.5, rely=0.2, anchor='center')
        self.entry_3.place(relx=0.5, rely=0.3, anchor='center')

        self.label_1 = CTkLabel(master=self.sub_pagina,width=20, text='R', font=('Arial', 12))
        self.label_2 = CTkLabel(master=self.sub_pagina,width=20, text='G', font=('Arial', 12))
        self.label_3 = CTkLabel(master=self.sub_pagina,width=20, text='B', font=('Arial', 12))

        self.label_1.place(relx=0.3, rely=0.1, anchor='center')
        self.label_2.place(relx=0.3, rely=0.2, anchor='center')
        self.label_3.place(relx=0.3, rely=0.3, anchor='center')

        self.btn_buscar = CTkButton(master=self.sub_pagina, text='Buscar', command= self.buscar_cor_RGB, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_buscar.place(relx=0.5, rely=0.8, anchor='center')

    def sub_page_2(self):
        self.sub_Pagina(170, 250)

        self.entry_1 = CTkEntry(master=self.sub_pagina, width=40, fg_color='transparent', text_color=cor_texto, font=('consolas', 12))
        self.entry_2 = CTkEntry(master=self.sub_pagina, width=40, fg_color='transparent', text_color=cor_texto, font=('consolas', 12))

        self.entry_1.place(relx=0.5, rely=0.1, anchor='center')
        self.entry_2.place(relx=0.5, rely=0.25, anchor='center')

        self.label_1 = CTkLabel(master=self.sub_pagina,width=20, text='X', font=('Arial', 12))
        self.label_2 = CTkLabel(master=self.sub_pagina,width=20, text='Y', font=('Arial', 12))

        self.label_1.place(relx=0.3, rely=0.1, anchor='center')
        self.label_2.place(relx=0.3, rely=0.25, anchor='center')


        self.btn_buscar = CTkButton(master=self.sub_pagina, text='Buscar', command=self.buscar_cor_posicao, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.btn_buscar.place(relx=0.5, rely=0.8, anchor='center')

    def sub_page_3(self):
        self.sub_Pagina(260, 220)

        self.text = CTkLabel(self.sub_pagina, text="Cor da imagem", font=('Arial', 13), width=130)
        self.text.place(relx=0.25, rely=0.1, anchor='center')

        self.text1 = CTkLabel(self.sub_pagina, text="Cor para trocar", font=('Arial', 13), width=130)
        self.text1.place(relx=0.75, rely=0.1, anchor='center')

        self.R_label_trocar = CTkLabel(self.sub_pagina, text="R", font=('Arial', 12), width=30)
        self.R_label_trocar.place(relx=0.1, rely=0.25, anchor='center')
        self.R_entry_trocar = CTkEntry(self.sub_pagina, font=('consolas', 12), width=40, fg_color='transparent', text_color=cor_texto)
        self.R_entry_trocar.place(relx=0.25, rely=0.25, anchor='center')

        self.G_label_trocar = CTkLabel(self.sub_pagina, text="G", font=('Arial', 12), width=30)
        self.G_label_trocar.place(relx=0.1, rely=0.45, anchor='center')
        self.G_entry_trocar = CTkEntry(self.sub_pagina, font=('consolas', 12), width=40, fg_color='transparent', text_color=cor_texto)
        self.G_entry_trocar.place(relx=0.25, rely=0.45, anchor='center')

        self.B_label_trocar = CTkLabel(self.sub_pagina, text="B", font=('Arial', 12), width=30)
        self.B_label_trocar.place(relx=0.1, rely=0.65, anchor='center')
        self.B_entry_trocar = CTkEntry(self.sub_pagina, font=('consolas', 12), width=40, fg_color='transparent', text_color=cor_texto)
        self.B_entry_trocar.place(relx=0.25, rely=0.65, anchor='center')


        self.R_label_trocar2 = CTkLabel(self.sub_pagina, text="R", font=('Arial', 12))
        self.R_label_trocar2.place(relx=0.6, rely=0.25, anchor='center')
        self.R_entry_trocar2 = CTkEntry(self.sub_pagina, font=('consolas', 12), width=40, fg_color='transparent', text_color=cor_texto)
        self.R_entry_trocar2.place(relx=0.75, rely=0.25, anchor='center')

        self.G_label_trocar2 = CTkLabel(self.sub_pagina, text="G", font=('Arial', 12))
        self.G_label_trocar2.place(relx=0.6, rely=0.45, anchor='center')
        self.G_entry_trocar2 = CTkEntry(self.sub_pagina, font=('consolas', 12), width=40, fg_color='transparent', text_color=cor_texto)
        self.G_entry_trocar2.place(relx=0.75, rely=0.45, anchor='center')

        self.B_label_trocar2 = CTkLabel(self.sub_pagina, text="B", font=('Arial', 12), width=30)
        self.B_label_trocar2.place(relx=0.6, rely=0.65, anchor='center')
        self.B_entry_trocar2 = CTkEntry(self.sub_pagina, font=('consolas', 12), width=40, fg_color='transparent', text_color=cor_texto)
        self.B_entry_trocar2.place(relx=0.75, rely=0.65, anchor='center')

        self.bt_sub_trocar_cor = CTkButton(self.sub_pagina, text='Trocar', command=self.trocar_cor, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.bt_sub_trocar_cor.place(relx= 0.5, rely= 0.85, anchor='center')

    def sub_page_4(self):
        self.sub_Pagina(160, 390)
        self.text = CTkLabel(self.sub_pagina, text="Cores usadas na imagem", font=('Arial', 12), width=130)
        self.text.place(relx=0.5, rely=0.1, anchor='center')
        self.listar_cores()

    def sub_page_5(self): 
        self.sub_Pagina(150, 150)

        self.text = CTkLabel(self.sub_pagina, text="Fator de blur", font=('Arial', 13), width=130)
        self.text.place(relx=0.5, rely=0.1, anchor='center')
        self.bt_sub_blur = CTkButton(self.sub_pagina, text='Blur 3', command=lambda: self.filtro_1(3), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.bt_sub_blur.place(relx= 0.5, rely= 0.35, anchor='center')
        self.bt_sub_blur = CTkButton(self.sub_pagina, text='Blur 5', command=lambda: self.filtro_1(5), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.bt_sub_blur.place(relx= 0.5, rely= 0.6, anchor='center')
        self.bt_sub_blur = CTkButton(self.sub_pagina, text='Blur 9', command=lambda: self.filtro_1(9), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        self.bt_sub_blur.place(relx= 0.5, rely= 0.85, anchor='center')
        
    def sub_page_6(self): 
        tipo = self.imagem.mode

        if tipo == "RGB":
            self.sub_Pagina(150, 150)
            
            self.text = CTkLabel(self.sub_pagina, text="Retirar canal de cor RGB", font=('Arial', 12))
            self.text.place(relx=0.5, rely=0.1, anchor='center')
            self.bt_sub_retirar_canal_R = CTkButton(self.sub_pagina, text='Retirar canal R', command=lambda: self.retirar_canal_de_cor(1, 'R'), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
            self.bt_sub_retirar_canal_R.place(relx= 0.5, rely= 0.35, anchor='center')
            self.bt_sub_retirar_canal_G = CTkButton(self.sub_pagina, text='Retirar canal G', command=lambda: self.retirar_canal_de_cor(1, 'G'), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
            self.bt_sub_retirar_canal_G.place(relx= 0.5, rely= 0.6, anchor='center')
            self.bt_sub_retirar_canal_B = CTkButton(self.sub_pagina, text='Retirar canal B', command=lambda: self.retirar_canal_de_cor(1, 'B'), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
            self.bt_sub_retirar_canal_B.place(relx= 0.5, rely= 0.85, anchor='center')

        elif tipo == "CMYK":
            self.sub_Pagina(150, 200)
            
            self.text = CTkLabel(self.sub_pagina, text="Retirar canal de cor CMYK", font=('Arial', 12))
            self.text.place(relx=0.5, rely=0.1, anchor='center')
            self.bt_retirar_canal_C = CTkButton(self.sub_pagina, text='Retirar canal C', command=lambda: self.retirar_canal_de_cor(2, 'C'), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
            self.bt_retirar_canal_C.place(relx= 0.5, rely= 0.25, anchor='center')
            self.bt_retirar_canal_M = CTkButton(self.sub_pagina, text='Retirar canal M', command=lambda: self.retirar_canal_de_cor(2, 'M'), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
            self.bt_retirar_canal_M.place(relx= 0.5, rely= 0.45, anchor='center')
            self.bt_retirar_canal_Y = CTkButton(self.sub_pagina, text='Retirar canal Y', command=lambda: self.retirar_canal_de_cor(2, 'Y'), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
            self.bt_retirar_canal_Y.place(relx= 0.5, rely= 0.65, anchor='center')
            self.bt_retirar_canal_K = CTkButton(self.sub_pagina, text='Retirar canal K', command=lambda: self.retirar_canal_de_cor(2, 'K'), 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
            self.bt_retirar_canal_K.place(relx= 0.5, rely= 0.85, anchor='center')



                                                                                        
                                                                                        
class classe_principal(Procetimentos, Funcoes):
    def __init__(self):
        self.root = root
        self.indi = 1 

        self.Tela_main()
        self.menu_principal()
        self.Butoes()

        root.mainloop()

    def Tela_main(self):
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width - largura_tela) // 2
        y = (screen_height - altura_tela) // 2
        root.geometry(f"{largura_tela}x{altura_tela}+{x}+{y}")
        root.resizable(False, False)
        root.configure(fg_color=cor_fundo, anchor='center')

    def menu_principal(self):
        self.root.title("Menu Principal")
        self.tabview  = CTkTabview(master=self.root, height=200, width=680, anchor='s', border_color='#0f160c', border_width=2,
                                   fg_color=cor_frame, segmented_button_fg_color=cor_1, segmented_button_selected_hover_color=cor_2,
                                   segmented_button_unselected_color='#738270',segmented_button_unselected_hover_color='#3b4838', 
                                   segmented_button_selected_color='#3b4838')
        self.tabview.place(relx=0.5, rely=0.85, anchor='center')

        self.tabview.add(" 1 ")
        self.tabview.add(" 2 ")
        self.tabview.add(" 3 ")
        self.tabview.add(" 4 ")
        self.tabview.add(" 5 ")
        self.tabview.add(" 6 ")
        self.tabview.add(" 7 ")
        self.tabview.add(" 8 ")
        self.tabview.add(" 9 ")

        self.titulo = CTkLabel(master=self.root, text="Processamento de Imagem", font=("castellar", 20), text_color='black')
        self.titulo.place(relx=0.5, rely=0.025, anchor='center')
        self.frame_1 = CTkFrame(master=self.root, height=400, width=550, fg_color=cor_frame)
        self.frame_1.place(relx=0.5, rely=0.37, anchor="center")
        self.frame_2 = CTkFrame(master=self.root, height=30, width=150, fg_color=cor_frame, corner_radius=0)
        self.frame_2.place(relx=0.5, rely=0.66, anchor="center")

        btn_abrir = CTkButton(self.frame_2, text="Abrir Imagem", command=self.abrir_imagem, 
                                      corner_radius=36, fg_color=cor_1, hover_color=cor_2)
        btn_abrir.place(relx=0.5, rely=0.5, anchor='center')


    def sub_Pagina(self, largura_pagina, altura_pagina):
        self.sub_pagina = Toplevel(master=self.root)
        self.sub_pagina.title("Retirar canal de cor")
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width - largura_pagina) // 2
        y = (screen_height - altura_pagina) // 3
        self.sub_pagina.geometry(f"{largura_pagina}x{altura_pagina}+{x}+{y}")
        self.sub_pagina.resizable(False, False)
        self.sub_pagina.configure(background=cor_frame)

    def sub_mensagem_de_erro(self):
        self.sub_mensagem_erro = Toplevel(master=self.root)
        self.sub_mensagem_erro.title("ERRO")
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width - 290) // 2
        y = (screen_height - 50) // 2
        self.sub_mensagem_erro.geometry(f"{290}x{50}+{x}+{y}")
        self.sub_mensagem_erro.resizable(False, False)
        self.sub_mensagem_erro.configure(background='dark red')

classe_principal()
